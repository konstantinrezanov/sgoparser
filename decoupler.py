import openpyxl
def decouple(clas, column):
    wb=openpyxl.load_workbook('init.xlsx')
    ws=wb.active
    a=0
    rows=[]
    for l in ['D', 'E', 'F', 'G', 'H']:
        for i in range(1, 41):
            if ws[l+str(i)].value==clas:
                a=i
                for n in range(a+1, a+9):
                    rows.append(ws[l+str(n)].value)
            else: 
                continue
    return \
    clas+'\n\n'+\
    '1  '+'8:00-8:40  '+str(rows[0])+ \
    '\n2  '+'8:50-9:30  '+str(rows[1])+ \
    '\n3  '+'9:50-10:30  '+ str(rows[2])+ \
    '\n4  '+'10:50-11:30  '+str(rows[3])+ \
    '\n5  '+'11:50-12:30  '+str(rows[4])+ \
    '\n6  '+'12:40-13:20  '+str(rows[5])+ \
    '\n7  '+'13:30-14:10  '+str(rows[6])+ \
    '\n8  '+'14:20-15:10  '+str(rows[7]) 